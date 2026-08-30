"""
Testes automatizados de ui/exportacao.py — exportação da carteira para
Excel (.xlsx) e CSV. Diferente dos outros arquivos de teste, este depende
de pandas e openpyxl (já são dependências do projeto, ver requirements.txt).

Rode com `pytest -v` (ver instruções em tests/test_calculations.py).
"""

from __future__ import annotations

import io

import openpyxl
import pytest

from ui import exportacao

DADOS_EXEMPLO = {
    "compras": [
        {"id": "1", "tipo": "compra", "ticker": "PETR4", "data": "2024-01-10", "qtd": 100, "preco": 30.0, "taxas": 5.0},
        {"id": "2", "tipo": "compra", "ticker": "VALE3", "data": "2024-02-15", "qtd": 50, "preco": 60.0, "taxas": 3.0},
        {"id": "3", "tipo": "venda", "ticker": "PETR4", "data": "2024-06-01", "qtd": 20, "preco": 35.0, "taxas": 2.0},
    ],
    "eventos": [],
    "cotacoes": {
        "PETR4": {"preco": 33.0, "previousClose": 32.0},
        "VALE3": {"preco": 58.0, "previousClose": 59.0},
    },
    "proventos": [
        {"data": "2024-03-10", "ticker": "PETR4", "tipo": "Dividendo", "valor": 45.5},
        {"data": "2024-07-10", "ticker": "VALE3", "tipo": "JCP", "valor": 20.0},
    ],
    "historico": [],
    "alertas": {},
    "setores": {"PETR4": "Petróleo e Gás", "VALE3": "Mineração e Siderurgia"},
    "precosTeto": {"PETR4": {"precoTeto": 40.0}},
    "watchlist": ["ITUB4"],
    "fundamentos": {},
    "exportadoEm": None,
}


def test_dataframe_posicoes_inclui_posicoes_reais_e_empresas_alvo():
    df = exportacao.montar_dataframe_posicoes(DADOS_EXEMPLO)
    assert len(df) == 3  # PETR4, VALE3 (posições) + ITUB4 (watchlist, alvo)
    assert set(df["Ticker"]) == {"PETR4", "VALE3", "ITUB4"}
    assert "Resultado (R$)" in df.columns


def test_dataframe_posicoes_vazio_quando_sem_nenhum_ativo():
    dados_vazios = {**DADOS_EXEMPLO, "compras": [], "watchlist": []}
    df = exportacao.montar_dataframe_posicoes(dados_vazios)
    assert len(df) == 0
    assert "Ticker" in df.columns  # cabeçalho existe mesmo vazio


def test_dataframe_proventos_ordena_mais_recente_primeiro():
    df = exportacao.montar_dataframe_proventos(DADOS_EXEMPLO)
    assert list(df["Data"]) == ["2024-07-10", "2024-03-10"]


def test_dataframe_transacoes_ordena_mais_recente_primeiro():
    df = exportacao.montar_dataframe_transacoes(DADOS_EXEMPLO)
    assert list(df["Data"]) == ["2024-06-01", "2024-02-15", "2024-01-10"]
    assert list(df["Tipo"]) == ["venda", "compra", "compra"]


def test_dataframe_resumo_ir_reflete_o_lucro_da_venda_registrada():
    df = exportacao.montar_dataframe_resumo_ir(DADOS_EXEMPLO)
    assert len(df) == 1
    linha = df.iloc[0]
    assert linha["Mês"] == "2024-06"
    # venda de 20 PETR4 a 35, preço médio de compra ~30.05 -> lucro positivo, pequeno o bastante pra isenção
    assert linha["Lucro Swing (R$)"] > 0
    assert linha["Isento (Swing)"] == "Sim"


def test_gerar_csv_posicoes_contem_cabecalho_e_dados():
    csv_texto = exportacao.gerar_csv_posicoes(DADOS_EXEMPLO)
    linhas = csv_texto.strip().splitlines()
    assert "Ticker" in linhas[0]
    assert len(linhas) == 4  # cabeçalho + 3 ativos


def test_gerar_excel_carteira_produz_arquivo_valido_com_4_abas():
    xlsx_bytes = exportacao.gerar_excel_carteira(DADOS_EXEMPLO)
    assert len(xlsx_bytes) > 1000

    pasta_trabalho = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert pasta_trabalho.sheetnames == ["Posições", "Proventos", "Compras e Vendas", "Resumo IR Mensal"]

    aba_posicoes = pasta_trabalho["Posições"]
    cabecalho = [c.value for c in next(aba_posicoes.iter_rows(min_row=1, max_row=1))]
    assert "Ticker" in cabecalho
    assert "Resultado (R$)" in cabecalho

    # confere que os dados batem, não só que o arquivo "abre"
    linhas_dados = list(aba_posicoes.iter_rows(min_row=2, values_only=True))
    tickers_na_planilha = {linha[cabecalho.index("Ticker")] for linha in linhas_dados}
    assert tickers_na_planilha == {"PETR4", "VALE3", "ITUB4"}


def test_gerar_excel_carteira_com_carteira_vazia_nao_quebra():
    dados_vazios = {**DADOS_EXEMPLO, "compras": [], "proventos": [], "watchlist": []}
    xlsx_bytes = exportacao.gerar_excel_carteira(dados_vazios)
    pasta_trabalho = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert pasta_trabalho.sheetnames == ["Posições", "Proventos", "Compras e Vendas", "Resumo IR Mensal"]
